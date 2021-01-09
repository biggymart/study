from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 Sheet를 기본 이름으로 생성
ws.title = "MySheet" # Sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # RGB 형태로 값을 넣어주면 탭 색상 변경 https://www.w3schools.com/colors/colors_rgb.asp

ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 Sheet 생성
ws2 = wb.create_sheet("NewSheet", 2) # 특정 위치에 넣고 싶은 경우 인덱스 사용

new_ws = wb["NewSheet"] # Dict 형태, 시트명으로 접근

print(wb.sheetnames) # 모든 시트명 확인 ['Sheet', 'MySheet', 'NewSheet', 'YourSheet']

# Sheet 복사
new_ws["A1"] = "Test" # A1 cell에 Test라는 값 입력
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")