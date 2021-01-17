from openpyxl import load_workbook
wb = load_workbook("sample.xlsx") # Load wb from sample.xlsx
ws = wb.active # Activated Sheet

# Objective: load and print cell data
for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end=" ") # 1 2 3 4 ...
    print()

# In case you don't know how many cells are in a sheet
for x in range(1, ws.max_row + 1):
    for y in range(1, ws.max_column + 1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()