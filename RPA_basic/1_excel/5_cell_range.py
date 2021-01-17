# Let's make a data
from openpyxl import Workbook
from random import *
wb = Workbook()
ws = wb.active

# Create a random data (e.g. School grade)
ws.append(["number", "English", "Math"]) # A, B, C
for i in range(1, 11):
    ws.append([i, randint(1, 100), randint(0, 100)])

# Suppose you are an English teacher and you don't need grades for math.
col_B = ws["B"]
# print(col_B) # (<Cell 'Sheet'.B1>, <Cell 'Sheet'.B2>, ...)
for cell in col_B:
    print(cell.value, end=" ")

print("\n======")

# If you don't need number:
col_range = ws["B:C"]
for cols in col_range:
    for cell in cols:
        print(cell.value, end=" ")
    print()
### format of nested loop ###
# for A in B:
#     for C in A:
#         do sth

print("\n======")

row_title = ws[1] # the first row, which is the title line
for cell in row_title:
    print(cell.value, end=" ")

print("\n======")

row_range = ws[2:6] # rows from 2 to 5 (students 1 to 5)
for rows in row_range:
    for cell in rows:
        print(cell.value, end=" ")
    print()

print("\n======")

row_range2 = ws[2:ws.max_row] # from row 2 to the end
for rows in row_range2:
    for cell in rows:
        print(cell.value, end=" ")
    print()

print("\n======")

# When you need cell location
from openpyxl.utils.cell import coordinate_from_string
row_range2 = ws[2:ws.max_row] # from row 2 to the end
for rows in row_range2:
    for cell in rows:
        print(cell.coordinate, end=" ") # A2 B2 C2
        xy = coordinate_from_string(cell.coordinate)
        print(xy, end=" ")  # ('A', '2') ('B', 2) ('C', 2) tuple
        print(xy[0], end=" ") # A B C
        print(xy[1], end=" ") # 2 2 2
    print()

wb.save("sample.xlsx")